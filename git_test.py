from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.cell(1,1).value = "Hello World!!!"
ws.cell(2,1).value = "GoodBuy World!!!"
ws.cell(3,1).value = "GitHub!!!"


wb.save('It\'s test.xlsx')
